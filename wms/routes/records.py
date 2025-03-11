from flask import render_template, request, redirect, url_for, send_file
from flask_login import login_required, current_user
from wms import app, db
from wms.utils import admin_required
from wms.models import (
    Receipt,
    ReceiptType,
    Warehouse,
    Area,
    Department,
    Transaction,
    ItemSKU,
    Item,
)
from datetime import datetime, timedelta
from sqlalchemy.orm import joinedload
from sqlalchemy import func, and_, select, distinct
from openpyxl import Workbook
from io import BytesIO
from decimal import Decimal


@app.route("/records", methods=["GET"])
@login_required
def records():
    # Get all unique item names for datalist
    item_names = (
        db.session.execute(select(distinct(Item.name)).order_by(Item.name))
        .scalars()
        .all()
    )

    # Get filter parameters from request
    record_type = request.args.get(
        "type", "stockout"
    )  # stockin or stockout (default to stockout)
    warehouse_id = request.args.get("warehouse")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    refcode = request.args.get("refcode")
    location_info = request.args.get("location_info")
    item_name = request.args.get("item_name")
    sku_desc = request.args.get("sku_desc")
    page = request.args.get("page", 1, type=int)
    per_page = 20

    # Get warehouses accessible by the current user
    if current_user.is_admin:
        warehouses = Warehouse.query.all()
    else:
        # Regular users can only see public warehouses and their own warehouse
        warehouses = Warehouse.query.filter(
            (Warehouse.is_public.is_(True)) | (Warehouse.owner_id == current_user.id)
        ).all()

    # For regular users, ensure a warehouse is selected
    if not current_user.is_admin and not warehouse_id and warehouses:
        # Default to user's own warehouse if available, otherwise first available warehouse
        user_warehouse = next(
            (w for w in warehouses if w.owner_id == current_user.id), None
        )
        default_warehouse = user_warehouse or warehouses[0] if warehouses else None

        if default_warehouse:
            # Redirect to same page with default warehouse selected
            return redirect(
                url_for(
                    "records",
                    warehouse=default_warehouse.id,
                    type=record_type,
                    start_date=start_date,
                    end_date=end_date,
                    refcode=refcode,
                    location_info=location_info,
                    item_name=item_name,
                    sku_desc=sku_desc,
                )
            )

    # Query base - join necessary relationships
    query = (
        db.session.query(Transaction)
        .join(Receipt)
        .join(Warehouse)
        .join(ItemSKU)
        .join(Item)
        .options(
            joinedload(Transaction.receipt).joinedload(Receipt.warehouse),
            joinedload(Transaction.receipt).joinedload(Receipt.operator),
            joinedload(Transaction.receipt).joinedload(Receipt.area),
            joinedload(Transaction.receipt).joinedload(Receipt.department),
            joinedload(Transaction.itemSKU).joinedload(ItemSKU.item),
        )
        .order_by(Receipt.date.desc(), Transaction.id.asc())
    )

    # Filter by user's warehouse access if not admin
    if not current_user.is_admin:
        query = query.filter(
            (Warehouse.is_public.is_(True)) | (Warehouse.owner_id == current_user.id)
        )

    # Apply filters
    if record_type == "stockin":
        query = query.filter(Receipt.type == ReceiptType.STOCKIN)
    else:
        query = query.filter(Receipt.type == ReceiptType.STOCKOUT)
        if location_info:
            # Search area or department name if provided
            query = (
                query.outerjoin(Area)
                .outerjoin(Department)
                .filter(
                    (Area.name.ilike(f"%{location_info}%"))
                    | (Department.name.ilike(f"%{location_info}%"))
                    | (Receipt.location.ilike(f"%{location_info}%"))
                )
            )

    if warehouse_id:
        # For non-admins, ensure they can only access their warehouse or public warehouses
        if not current_user.is_admin:
            allowed_warehouse_ids = [w.id for w in warehouses]
            if int(warehouse_id) not in allowed_warehouse_ids:
                warehouse_id = None

        if warehouse_id:
            query = query.filter(Receipt.warehouse_id == warehouse_id)

    if start_date:
        start_datetime = datetime.strptime(
            f"{start_date} 00:00:00", "%Y-%m-%d %H:%M:%S"
        )
        query = query.filter(Receipt.date >= start_datetime)

    if end_date:
        end_datetime = datetime.strptime(f"{end_date} 23:59:59", "%Y-%m-%d %H:%M:%S")
        query = query.filter(Receipt.date <= end_datetime)

    if refcode and record_type == "stockin":
        query = query.filter(Receipt.refcode.ilike(f"%{refcode}%"))

    # Add new filters for item name and SKU description
    if item_name:
        query = query.filter(Item.name.ilike(f"%{item_name}%"))

    if sku_desc:
        query = query.filter(
            (ItemSKU.brand.ilike(f"%{sku_desc}%"))
            | (ItemSKU.spec.ilike(f"%{sku_desc}%"))
        )

    # Paginate results - now based on transactions
    pagination = query.paginate(page=page, per_page=per_page)

    return render_template(
        "records.html.jinja",
        pagination=pagination,
        warehouses=warehouses,
        record_type=record_type,
        warehouse_id=warehouse_id,
        start_date=start_date,
        end_date=end_date,
        refcode=refcode,
        location_info=location_info,
        item_name=item_name,
        sku_desc=sku_desc,
        item_names=item_names,
        request=request,
    )


@app.route("/statistics_fee", methods=["GET"])
@login_required
@admin_required
def statistics_fee():
    # Get current year and month for default date range
    today = datetime.now()
    current_year = today.year
    current_month = today.month

    # Set default date range to current month if not provided
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")

    # If no dates are provided, default to current month
    if not start_date and not end_date:
        # First day of current month
        start_date = f"{current_year}-{current_month:02d}-01"

        # Last day of current month - calculate based on the next month's first day minus one day
        if current_month == 12:
            next_month_year = current_year + 1
            next_month = 1
        else:
            next_month_year = current_year
            next_month = current_month + 1

        # Create first day of next month
        next_month_first = datetime(next_month_year, next_month, 1)
        # Subtract one day to get last day of current month
        last_day = (next_month_first - timedelta(days=1)).day
        end_date = f"{current_year}-{current_month:02d}-{last_day}"

    # Initialize empty data structures
    warehouses = Warehouse.query.order_by(Warehouse.name).all()
    areas = Area.query.order_by(Area.name).all()
    departments = Department.query.order_by(Department.name).all()

    # Initialize statistics data structure with Decimal(0) instead of 0
    stats_data = {
        "warehouses": {w.id: {"name": w.name} for w in warehouses},
        "areas": {a.id: {"name": a.name, "departments": {}, "total": Decimal('0')} for a in areas},
        "departments": {d.id: {"name": d.name, "total": Decimal('0')} for d in departments},
        "total_by_warehouse": {w.id: Decimal('0') for w in warehouses},
        "grand_total": Decimal('0'),
    }

    # Initialize each cell in the warehouse × area × department structure
    for warehouse in warehouses:
        stats_data["warehouses"][warehouse.id]["areas"] = {}
        for area in areas:
            stats_data["warehouses"][warehouse.id]["areas"][area.id] = {
                "total": 0,
                "departments": {},
            }
            for department in departments:
                stats_data["warehouses"][warehouse.id]["areas"][area.id]["departments"][
                    department.id
                ] = 0

    # Process filter date range
    filter_conditions = [Receipt.type == ReceiptType.STOCKOUT]

    if start_date:
        start_datetime = datetime.strptime(
            f"{start_date} 00:00:00", "%Y-%m-%d %H:%M:%S"
        )
        filter_conditions.append(Receipt.date >= start_datetime)

    if end_date:
        end_datetime = datetime.strptime(f"{end_date} 23:59:59", "%Y-%m-%d %H:%M:%S")
        filter_conditions.append(Receipt.date <= end_datetime)

    # Query aggregating transactions by warehouse, area, and department
    results = (
        db.session.query(
            Receipt.warehouse_id,
            Receipt.area_id,
            Receipt.department_id,
            func.sum(Transaction.count * Transaction.price * Decimal('-1')).label("total_value"),
        )
        .join(Transaction)
        .filter(and_(*filter_conditions))
        .group_by(Receipt.warehouse_id, Receipt.area_id, Receipt.department_id)
        .all()
    )

    # Populate the statistics data structure
    for warehouse_id, area_id, department_id, total_value in results:
        # Skip entries with no area or department
        if not area_id or not department_id:
            continue

        # Keep as Decimal for precision
        value = total_value

        # Update values in the nested structure
        stats_data["warehouses"][warehouse_id]["areas"][area_id]["departments"][
            department_id
        ] = value

        # Update area total for this warehouse
        stats_data["warehouses"][warehouse_id]["areas"][area_id]["total"] += value

        # Update warehouse total
        stats_data["total_by_warehouse"][warehouse_id] += value

        # Update area total
        stats_data["areas"][area_id]["total"] += value

        # Update department total
        stats_data["departments"][department_id]["total"] += value

        # Update area's department structure (for the area tab)
        if department_id not in stats_data["areas"][area_id]["departments"]:
            stats_data["areas"][area_id]["departments"][department_id] = Decimal('0')
        stats_data["areas"][area_id]["departments"][department_id] += value

        # Update grand total
        stats_data["grand_total"] += value

    return render_template(
        "statistics_fee.html.jinja",
        start_date=start_date,
        end_date=end_date,
        warehouses=warehouses,
        areas=areas,
        departments=departments,
        stats_data=stats_data,
        current_year=current_year,
        current_month=current_month,
    )


@app.route("/statistics_usage", methods=["GET"])
@login_required
def statistics_usage():
    # Get current year and month for default date range
    today = datetime.now()
    current_year = today.year
    current_month = today.month

    # Get filter parameters from request
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    warehouse_id = request.args.get("warehouse")
    item_name = request.args.get("item_name", "")
    brand = request.args.get("brand", "")
    spec = request.args.get("spec", "")

    # If no dates are provided, default to current month
    if not start_date and not end_date:
        # First day of current month
        start_date = f"{current_year}-{current_month:02d}-01"

        # Last day of current month
        if current_month == 12:
            next_month_year = current_year + 1
            next_month = 1
        else:
            next_month_year = current_year
            next_month = current_month + 1
        next_month_first = datetime(next_month_year, next_month, 1)
        last_day = (next_month_first - timedelta(days=1)).day
        end_date = f"{current_year}-{current_month:02d}-{last_day}"

    # Get warehouses accessible by the current user
    if current_user.is_admin:
        warehouses = Warehouse.query.all()
    else:
        warehouses = Warehouse.query.filter(
            (Warehouse.is_public.is_(True)) | (Warehouse.owner_id == current_user.id)
        ).all()

    # For regular users, ensure a warehouse is selected
    if not current_user.is_admin and not warehouse_id and warehouses:
        user_warehouse = next(
            (w for w in warehouses if w.owner_id == current_user.id), None
        )
        default_warehouse = user_warehouse or warehouses[0] if warehouses else None
        if default_warehouse:
            return redirect(
                url_for(
                    "statistics_usage",
                    warehouse=default_warehouse.id,
                    start_date=start_date,
                    end_date=end_date,
                )
            )

    # Get all unique item names for datalist
    item_names = db.session.execute(select(distinct(Item.name)).order_by(Item.name)).scalars().all()

    # Query base - aggregate transactions by ItemSKU
    query = (
        db.session.query(
            ItemSKU,
            Item,
            func.sum(Transaction.count * -1).label("total_usage"),
        )
        .join(Transaction)
        .join(Receipt)
        .join(Item)
        .filter(Receipt.type == ReceiptType.STOCKOUT)
        .group_by(ItemSKU.id, Item.id)
        .order_by(Item.name, ItemSKU.brand, ItemSKU.spec)
    )

    # Filter by warehouse access if not admin
    if not current_user.is_admin:
        query = query.join(Receipt.warehouse).filter(
            (Warehouse.is_public.is_(True)) | (Warehouse.owner_id == current_user.id)
        )

    # Apply filters
    if warehouse_id:
        if not current_user.is_admin:
            allowed_warehouse_ids = [w.id for w in warehouses]
            if int(warehouse_id) not in allowed_warehouse_ids:
                warehouse_id = None

        if warehouse_id:
            query = query.filter(Receipt.warehouse_id == warehouse_id)

    if start_date:
        start_datetime = datetime.strptime(
            f"{start_date} 00:00:00", "%Y-%m-%d %H:%M:%S"
        )
        query = query.filter(Receipt.date >= start_datetime)

    if end_date:
        end_datetime = datetime.strptime(f"{end_date} 23:59:59", "%Y-%m-%d %H:%M:%S")
        query = query.filter(Receipt.date <= end_datetime)

    # Apply item filters
    if item_name:
        query = query.filter(Item.name.ilike(f"%{item_name}%"))
    if brand:
        query = query.filter(ItemSKU.brand.ilike(f"%{brand}%"))
    if spec:
        query = query.filter(ItemSKU.spec.ilike(f"%{spec}%"))

    # Get results
    usage_data = query.all()

    return render_template(
        "statistics_usage.html.jinja",
        warehouses=warehouses,
        warehouse_id=warehouse_id,
        start_date=start_date,
        end_date=end_date,
        usage_data=usage_data,
        current_year=current_year,
        current_month=current_month,
        item_names=item_names,
        item_name=item_name,
        brand=brand,
        spec=spec,
    )


@app.route("/export_records")
@login_required
def export_records():
    # Get filter parameters from request - same as records route
    record_type = request.args.get("type", "stockout")
    warehouse_id = request.args.get("warehouse")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    refcode = request.args.get("refcode")
    location_info = request.args.get("location_info")
    item_name = request.args.get("item_name")
    sku_desc = request.args.get("sku_desc")

    # Get warehouses accessible by the current user - same as records route
    if current_user.is_admin:
        warehouses = Warehouse.query.all()
    else:
        warehouses = Warehouse.query.filter(
            (Warehouse.is_public.is_(True)) | (Warehouse.owner_id == current_user.id)
        ).all()

    # Query base - join necessary relationships
    query = (
        db.session.query(Transaction)
        .join(Receipt)
        .join(Warehouse)
        .join(ItemSKU)
        .join(Item)
        .options(
            joinedload(Transaction.receipt).joinedload(Receipt.warehouse),
            joinedload(Transaction.receipt).joinedload(Receipt.operator),
            joinedload(Transaction.receipt).joinedload(Receipt.area),
            joinedload(Transaction.receipt).joinedload(Receipt.department),
            joinedload(Transaction.itemSKU).joinedload(ItemSKU.item),
        )
        .order_by(Receipt.date.desc(), Transaction.id.asc())
    )

    # Apply all the same filters as the records route
    if not current_user.is_admin:
        query = query.filter(
            (Warehouse.is_public.is_(True)) | (Warehouse.owner_id == current_user.id)
        )

    if record_type == "stockin":
        query = query.filter(Receipt.type == ReceiptType.STOCKIN)
    else:
        query = query.filter(Receipt.type == ReceiptType.STOCKOUT)
        if location_info:
            query = (
                query.outerjoin(Area)
                .outerjoin(Department)
                .filter(
                    (Area.name.ilike(f"%{location_info}%"))
                    | (Department.name.ilike(f"%{location_info}%"))
                    | (Receipt.location.ilike(f"%{location_info}%"))
                )
            )

    if warehouse_id:
        if not current_user.is_admin:
            allowed_warehouse_ids = [w.id for w in warehouses]
            if int(warehouse_id) not in allowed_warehouse_ids:
                warehouse_id = None

        if warehouse_id:
            query = query.filter(Receipt.warehouse_id == warehouse_id)

    if start_date:
        start_datetime = datetime.strptime(
            f"{start_date} 00:00:00", "%Y-%m-%d %H:%M:%S"
        )
        query = query.filter(Receipt.date >= start_datetime)

    if end_date:
        end_datetime = datetime.strptime(f"{end_date} 23:59:59", "%Y-%m-%d %H:%M:%S")
        query = query.filter(Receipt.date <= end_datetime)

    if refcode and record_type == "stockin":
        query = query.filter(Receipt.refcode.ilike(f"%{refcode}%"))

    if item_name:
        query = query.filter(Item.name.ilike(f"%{item_name}%"))

    if sku_desc:
        query = query.filter(
            (ItemSKU.brand.ilike(f"%{sku_desc}%"))
            | (ItemSKU.spec.ilike(f"%{sku_desc}%"))
        )

    # Get all results for export
    transactions = query.all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "操作记录"

    # Write headers
    headers = ["日期", "物品", "规格", "数量", "价格", "仓库"]
    if record_type == "stockin":
        headers.append("单号")
    else:
        headers.extend(["操作员", "区域", "部门", "具体地点"])

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row, trans in enumerate(transactions, start=2):
        ws.cell(row=row, column=1, value=trans.receipt.date.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=2, value=trans.itemSKU.item.name)
        ws.cell(
            row=row, column=3, value=f"{trans.itemSKU.brand} - {trans.itemSKU.spec}"
        )
        ws.cell(
            row=row,
            column=4,
            value=trans.count if record_type == "stockin" else -trans.count,
        )
        # Format price as string with 2 decimal places for Excel
        ws.cell(row=row, column=5, value="{:.2f}".format(float(trans.price)))
        ws.cell(row=row, column=6, value=trans.receipt.warehouse.name)

        if record_type == "stockin":
            ws.cell(row=row, column=7, value=trans.receipt.refcode)
        else:
            ws.cell(row=row, column=7, value=trans.receipt.operator.nickname)
            area_name = trans.receipt.area.name if trans.receipt.area else ""
            ws.cell(row=row, column=8, value=area_name)
            department_name = (
                trans.receipt.department.name if trans.receipt.department else ""
            )
            ws.cell(row=row, column=9, value=department_name)
            location = trans.receipt.location or ""
            ws.cell(row=row, column=10, value=location)

    # Format price column as number with 2 decimal places
    for row in range(2, len(transactions) + 2):
        cell = ws.cell(row=row, column=5)
        cell.number_format = '0.00'

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Get warehouse name for filename
    warehouse_name = "全部仓库"
    if warehouse_id:
        warehouse = next((w for w in warehouses if w.id == int(warehouse_id)), None)
        if warehouse:
            warehouse_name = warehouse.name

    # Generate filename based on current datetime and warehouse
    filename = (
        f"records_{warehouse_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
